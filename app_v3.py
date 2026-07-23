# Validador de Leads v3 — Soluções Industriais
# v2 (busca automática no Metabase) + dashboard HTML de resultados para download.

import base64
import csv
import io
import json
import os
import re
import time
from datetime import date, datetime, timedelta

import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def svg_logo_si(cor, fundo):
    """Marca 'Soluções Industriais' recriada em SVG (skyline + smartphone).
    Não há como usar o arquivo de imagem original diretamente aqui — então
    recriamos o ícone, colorido conforme o tema (cor do traço + cor de fundo
    do 'recorte' da tela do aparelho)."""
    return f"""<svg class="logo-si" width="28" height="17" viewBox="0 0 300 180" xmlns="http://www.w3.org/2000/svg">
  <path d="M0 150 L40 95 L60 118 L92 72 L112 100 L132 62 L150 100 L150 150 Z" fill="{cor}"/>
  <path d="M300 150 L260 95 L240 118 L208 72 L188 100 L168 62 L150 100 L150 150 Z" fill="{cor}"/>
  <rect x="16" y="126" width="13" height="10" fill="{cor}"/>
  <rect x="58" y="126" width="13" height="10" fill="{cor}"/>
  <rect x="229" y="126" width="13" height="10" fill="{cor}"/>
  <rect x="271" y="126" width="13" height="10" fill="{cor}"/>
  <path d="M228 60 q10 -8 4 -20" stroke="{cor}" stroke-width="4" fill="none" stroke-linecap="round"/>
  <rect x="112" y="16" width="76" height="140" rx="20" fill="{cor}"/>
  <rect x="126" y="30" width="48" height="94" rx="7" fill="{fundo}"/>
  <circle cx="150" cy="62" r="13" fill="{fundo}"/>
  <circle cx="150" cy="62" r="5.5" fill="{cor}"/>
  <circle cx="150" cy="146" r="6" fill="{fundo}"/>
</svg>"""


# Cores de destaque no Excel por status
FILL_DENTRO = PatternFill("solid", fgColor="C6EFCE")   # verde
FILL_FORA = PatternFill("solid", fgColor="FFC7CE")     # vermelho
FILL_ABERTO = PatternFill("solid", fgColor="FFEB9C")   # âmbar
FILL_CABECALHO = PatternFill("solid", fgColor="0C447C")  # azul escuro

# Provedores gratuitos (APIs compatíveis com OpenAI). O app escolhe automaticamente:
# se houver CEREBRAS_API_KEY nos Secrets usa Cerebras (cota diária bem maior);
# senão usa Groq. Um serve de reserva do outro quando ambas as chaves existem.
# Cada provedor tem uma LISTA de modelos candidatos. Se um estiver descontinuado
# (404), o app tenta o próximo automaticamente — nunca quebra por um nome só.
# Modelos confirmados ATIVOS (consultados na doc oficial de cada provedor, jul/2026).
# Cada provedor tem uma lista de candidatos: se um cair (404), tenta o próximo.
PROVEDORES = {
    "cerebras": {
        "url": "https://api.cerebras.ai/v1/chat/completions",
        "chave": "CEREBRAS_API_KEY",
        "modelos": ["llama-3.3-70b", "llama3.1-8b", "gpt-oss-120b"],
    },
    "groq": {
        "url": "https://api.groq.com/openai/v1/chat/completions",
        "chave": "GROQ_API_KEY",
        "modelos": ["openai/gpt-oss-20b", "llama-3.1-8b-instant",
                    "llama-3.3-70b-versatile", "openai/gpt-oss-120b"],
    },
}

# Opções do seletor de IA na tela. Cada opção define (provedor, modelo preferido).
# Todos os modelos abaixo foram confirmados ativos na documentação oficial.
# Cerebras tem cota diária muito maior (1M tokens/dia) — melhor para uso intenso.
MODELOS_ESCOLHA = {
    "Automático (recomendado)": None,
    "Rápido · Cerebras Llama 3.1 8B": ("cerebras", "llama3.1-8b"),
    "Rápido · Groq GPT-OSS 20B": ("groq", "openai/gpt-oss-20b"),
    "Equilíbrio · Cerebras Llama 3.3 70B": ("cerebras", "llama-3.3-70b"),
    "Equilíbrio · Groq Llama 3.3 70B": ("groq", "llama-3.3-70b-versatile"),
    "Preciso · Groq GPT-OSS 120B": ("groq", "openai/gpt-oss-120b"),
}

TAMANHO_LOTE = 20
MAX_TENTATIVAS = 5
LIMITE_FONTE = 4000
LIMITE_PERFIL = 9000
STATUS_VALIDOS = {"Dentro do foco", "Fora do foco", "Aberto"}

CARD_ORCAMENTOS = 47   # question do Metabase: base de orçamentos por chave única
CARD_BRIEFING = 286    # question do Metabase: briefing do cliente por chave única
CARD_ANUNCIOS = 185    # question do Metabase: anúncios por chave única

AZUL = "#185FA5"
AZUL_ESCURO = "#0C447C"
AZUL_CLARO = "#E6F1FB"

ARQUIVO_HISTORICO = "historico.json"
PASTA_RESULTADOS = "resultados"

PROMPT_SISTEMA = """Você é um analista de qualidade de leads de uma plataforma de geração de leads B2B.
Sua tarefa: para cada lead abaixo, decidir se ele está DENTRO ou FORA do foco do cliente descrito no perfil, ou se está ABERTO (mensagem sem informação suficiente para avaliar).

ANTES DE CLASSIFICAR: identifique no perfil O QUE o cliente vende — um PRODUTO (ex. máquinas, equipamentos) ou um SERVIÇO (ex. corte sob medida, usinagem para terceiros). Essa distinção é o critério mais importante.

Critérios (avalie em conjunto, nenhum sozinho decide):
1. Produto vs. serviço. Se o cliente VENDE MÁQUINAS e o lead quer CONTRATAR o serviço (ex. "preciso cortar 50 chapas", "orçamento para corte de peças"), é "Fora do foco" — mesmo que a mensagem seja tecnicamente detalhada (material, medidas, CNPJ). Especificidade técnica NÃO transforma um pedido de serviço em lead de máquina. O inverso também vale (cliente presta serviço e lead quer comprar máquina). Salvo se o perfil disser que o cliente atende ambos.
2. Modalidade compatível. Pedidos de assistência técnica/manutenção, aluguel de máquina, ou peças/componentes avulsos são "Fora do foco" quando o cliente vende equipamentos novos — salvo indicação contrária no perfil ou nas regras específicas.
3. Material/produto compatível com o portfólio do cliente. Se o cliente trabalha metal e o lead pede madeira/tecido/PVC, é forte sinal de "Fora do foco", mesmo que o serviço seja o mesmo.
4. B2B vs. uso pessoal. Pedidos claramente domésticos/pontuais de pessoa física pesam para "Fora do foco" quando o cliente atende indústria/B2B.
5. Especificidade técnica. Medidas, normas, quantidade definida, nome de empresa/CNPJ pesam para "Dentro do foco" — mas SOMENTE quando o pedido é da modalidade certa (ver critério 1).
6. Sinais de ruído. Teste interno (QA, e-mails de qualidade), spam, concorrente se oferecendo, marca/modelo que o cliente não vende, ou lead avisando que já comprou em outro lugar = "Fora do foco" independente do produto.
7. Mensagem inteiramente em inglês. Se a mensagem do lead estiver totalmente escrita em inglês (ex. "Dear Sir/Madam, we are interested in your products..."), classifique como "Fora do foco" — são tipicamente bots ou contatos genéricos internacionais fora do público-alvo. Isso vale mesmo que a mensagem pareça pedir um produto do cliente. NÃO se aplica a mensagens em português que contenham apenas termos técnicos ou nomes de produto em inglês (ex. "máquina laser CO2", "new laser nli390") — essas continuam sendo avaliadas normalmente.
8. Regras específicas do cliente (se fornecidas no perfil) têm prioridade sobre os critérios gerais.

Regras de saída:
- STATUS deve ser EXATAMENTE um destes: "Dentro do foco", "Fora do foco", "Aberto".
- MOTIVO: uma frase objetiva em português citando a evidência da própria mensagem. O motivo deve justificar o STATUS escolhido, não outro.
- Mensagens vagas demais para julgar (ex. apenas "aço inox", apenas "me manda o e-mail", uma palavra solta sem contexto de compra) = "Aberto". NUNCA marque "Dentro do foco" sem evidência de interesse na modalidade certa (compra do produto que o cliente vende).
- Peças, componentes e insumos avulsos (ex. fonte, tubo de laser, lentes) = "Fora do foco" quando o cliente vende máquinas completas, salvo indicação contrária no perfil.
- Mensagens idênticas ou quase idênticas (mesmo texto em vários leads) DEVEM receber exatamente a mesma classificação e o mesmo motivo — revise antes de responder.
- Mensagem inteiramente em inglês = "Fora do foco" (ver critério 7), com motivo indicando que é mensagem em inglês / provável bot.
- Responda SOMENTE com um objeto JSON: {"resultados": [{"id": "...", "status": "...", "motivo": "..."}]} — um item por lead, na mesma ordem."""


# ---------- Histórico ----------

def carregar_historico():
    try:
        with open(ARQUIVO_HISTORICO, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


LIMITE_HISTORICO = 5  # guarda só as últimas consultas, para não acumular disco/memória


def salvar_no_historico(registro, xlsx_bytes=None, dash_bytes=None):
    historico = carregar_historico()
    historico.insert(0, registro)
    mantidos = historico[:LIMITE_HISTORICO]
    removidos = historico[LIMITE_HISTORICO:]
    try:
        with open(ARQUIVO_HISTORICO, "w", encoding="utf-8") as f:
            json.dump(mantidos, f, ensure_ascii=False)
        os.makedirs(PASTA_RESULTADOS, exist_ok=True)
        rid = registro.get("id", "")
        if rid and xlsx_bytes:
            with open(os.path.join(PASTA_RESULTADOS, f"{rid}.xlsx"), "wb") as f:
                f.write(xlsx_bytes)
        if rid and dash_bytes:
            with open(os.path.join(PASTA_RESULTADOS, f"{rid}.html"), "wb") as f:
                f.write(dash_bytes)
        # limpa os arquivos das consultas que saíram do histórico
        for antigo in removidos:
            rid_antigo = antigo.get("id", "")
            if not rid_antigo:
                continue
            for ext in (".csv", ".xlsx", ".html"):
                caminho = os.path.join(PASTA_RESULTADOS, f"{rid_antigo}{ext}")
                if os.path.exists(caminho):
                    os.remove(caminho)
    except Exception:
        pass


def excluir_do_historico(rid):
    historico = [h for h in carregar_historico() if h.get("id") != rid]
    try:
        with open(ARQUIVO_HISTORICO, "w", encoding="utf-8") as f:
            json.dump(historico, f, ensure_ascii=False)
        for ext in (".csv", ".xlsx", ".html"):
            caminho = os.path.join(PASTA_RESULTADOS, f"{rid}{ext}")
            if os.path.exists(caminho):
                os.remove(caminho)
    except Exception:
        pass


def ler_resultado_salvo(rid, ext):
    try:
        with open(os.path.join(PASTA_RESULTADOS, f"{rid}{ext}"), "rb") as f:
            return f.read()
    except Exception:
        return None


# ---------- Metabase ----------

def secret(nome, padrao=""):
    try:
        return str(st.secrets.get(nome, padrao)).strip()
    except Exception:
        return os.environ.get(nome, padrao).strip()


def cabecalhos_metabase():
    """Autentica no Metabase por API key ou por login/senha (sessão)."""
    api_key = secret("METABASE_API_KEY")
    if api_key:
        return {"X-API-KEY": api_key}
    if "mb_sessao" in st.session_state:
        return {"X-Metabase-Session": st.session_state["mb_sessao"]}
    usuario, senha = secret("METABASE_USER"), secret("METABASE_PASSWORD")
    if not (usuario and senha):
        return None
    r = requests.post(
        secret("METABASE_URL").rstrip("/") + "/api/session",
        json={"username": usuario, "password": senha},
        timeout=30,
    )
    r.raise_for_status()
    st.session_state["mb_sessao"] = r.json()["id"]
    return {"X-Metabase-Session": st.session_state["mb_sessao"]}


def consultar_question(card_id, parametros):
    """Baixa o resultado CSV de uma question do Metabase com parâmetros."""
    url = secret("METABASE_URL").rstrip("/") + f"/api/card/{card_id}/query/csv"
    headers = cabecalhos_metabase()
    if headers is None:
        raise RuntimeError(
            "Credenciais do Metabase não configuradas. Adicione nos Secrets: "
            "METABASE_URL e (METABASE_API_KEY ou METABASE_USER + METABASE_PASSWORD)."
        )
    params_mb = [
        {"type": tipo, "target": ["variable", ["template-tag", nome]], "value": valor}
        for nome, valor, tipo in parametros
    ]
    r = requests.post(url, headers=headers, data={"parameters": json.dumps(params_mb)}, timeout=120)
    if r.status_code == 401 and "mb_sessao" in st.session_state:
        del st.session_state["mb_sessao"]           # sessão expirada → renova e tenta de novo
        headers = cabecalhos_metabase()
        r = requests.post(url, headers=headers, data={"parameters": json.dumps(params_mb)}, timeout=120)
    r.raise_for_status()
    return r.content.decode("utf-8-sig", errors="replace")


# ---------- Fontes auxiliares ----------

def csv_anuncios_para_texto(conteudo_csv, limite=LIMITE_FONTE):
    """Transforma o CSV de anúncios em uma lista compacta de nomes."""
    linhas = list(csv.reader(io.StringIO(conteudo_csv)))
    if len(linhas) < 2:
        return ""
    itens = []
    for r in linhas[1:]:
        valores = [c.strip() for c in r if c.strip()]
        if valores:
            itens.append(" | ".join(valores))
    return "\n".join(itens)[:limite]


def extrair_nome_empresa(conteudo_csv):
    """Procura o nome do cliente em qualquer CSV do Metabase que tenha essa coluna."""
    try:
        linhas = list(csv.reader(io.StringIO(conteudo_csv)))
        if len(linhas) < 2:
            return None
        h = [c.strip().lower() for c in linhas[0]]
        for cand in ("nome fantasia", "nome da empresa", "empresa"):
            if cand in h:
                v = linhas[1][h.index(cand)].strip()
                if v:
                    return v
    except Exception:
        pass
    return None


def buscar_site(url, limite=LIMITE_FONTE):
    try:
        r = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        html = r.text
    except Exception:
        return ""
    html = re.sub(r"(?is)<(script|style|noscript)[^>]*>.*?</\1>", " ", html)
    texto = re.sub(r"(?s)<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", texto).strip()[:limite]


def csv_briefing_para_texto(conteudo_csv, limite=LIMITE_FONTE):
    """Converte o CSV do briefing (colunas longas) em texto legível para a IA."""
    linhas = list(csv.reader(io.StringIO(conteudo_csv)))
    if len(linhas) < 2:
        return ""
    h = linhas[0]
    blocos = []
    for r in linhas[1:]:
        campos = [f"{h[j]}: {r[j].strip()}" for j in range(min(len(h), len(r))) if r[j].strip()]
        blocos.append("\n".join(campos))
    return ("\n\n--- produto/linha seguinte ---\n\n".join(blocos))[:limite]


# ---------- Dashboard HTML ----------

MODELO_DASH = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Relatório de Validação — __EMPRESA__</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: -apple-system, 'Segoe UI', Arial, sans-serif; background: __BG_PAGINA__; }
  .tela { max-width: 1180px; margin: 0 auto; background: __BG_TELA__; }
  .nav { display: flex; align-items: center; justify-content: space-between; padding: 18px 36px; background: __BG_NAV__; border-bottom: 1px solid __BORDA_NAV__; }
  .nav .logo { display: flex; align-items: center; gap: 10px; color: __TXT_LOGO__; font-weight: 700; font-size: 14px; }
  .nav .logo .logo-si { flex-shrink: 0; }
  .nav .dir { display: flex; align-items: center; gap: 10px; }
  .nav .baixar { font-size: 12px; color: __BAIXAR_TXT__; background: __BAIXAR_BG__; padding: 8px 16px; border-radius: 20px; font-weight: 700; text-decoration: none; border: none; cursor: pointer; }
  .nav .baixar-imagem { background: transparent; border: 1px solid __SELO_BORDA__; color: __TXT_LOGO__; }
  .hero {
    padding: 24px 36px 20px;
    background: __HERO_OVERLAY__,
      url('https://images.unsplash.com/photo-1513828583688-c52646db42da?w=1900&q=70') center/cover no-repeat;
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px;
  }
  .hero h1 { color: __TXT_TITULO__; font-size: 20px; font-weight: 700; margin-bottom: 4px; }
  .hero p { color: __TXT_SUB__; font-size: 12.5px; }
  .selo { background: __SELO_BG__; border: 1px solid __SELO_BORDA__; color: __SELO_TXT__; font-size: 12.5px; padding: 7px 16px; border-radius: 30px; }
  .selo b { font-size: 14px; }
  .corpo { padding: 24px 36px 30px; }
  .kpis { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 18px; }
  .kpi { background: __BG_KPI__; border: 1px solid __BORDA_KPI__; border-radius: 14px; padding: 16px 18px; }
  .kpi .lbl { font-size: 11px; color: __TXT_LBL__; font-weight: 600; text-transform: uppercase; letter-spacing: .4px; margin-bottom: 6px; }
  .kpi .num { font-size: 24px; font-weight: 700; color: __TXT_NUM__; }
  .kpi .num.v { color: #1D9E75; } .kpi .num.r { color: #D85A30; } .kpi .num.a { color: #BA7517; }
  .kpi .sub { font-size: 11px; color: __TXT_LBL__; margin-top: 4px; }
  .linha { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 14px; }
  .painel { background: __BG_KPI__; border: 1px solid __BORDA_KPI__; border-radius: 14px; padding: 20px 22px; }
  .painel h2 { color: __TXT_NUM__; font-size: 13.5px; font-weight: 700; margin-bottom: 12px; }
  .painel h2 .pin { font-size: 11px; padding: 2px 8px; border-radius: 20px; vertical-align: middle; margin-left: 6px; font-weight: 600; }
  .pin-verde { background: rgba(29,158,117,0.18); color: #1D9E75; }
  .pin-verm { background: rgba(216,90,48,0.18); color: #D85A30; }
  .grafico { height: 210px; position: relative; }
  .lista-lead { display: flex; align-items: center; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid __BORDA_LISTA__; font-size: 12.5px; gap: 10px; }
  .lista-lead:last-child { border-bottom: none; }
  .lista-lead .nome { color: __TXT_NOME_LEAD__; font-weight: 600; display: flex; align-items: center; gap: 8px; }
  .pin-dot { width: 7px; height: 7px; border-radius: 50%; flex-shrink: 0; }
  .pin-dot-verde { background: #1D9E75; } .pin-dot-verm { background: #D85A30; }
  .lista-lead .email { color: __TXT_EMAIL_LEAD__; font-size: 11.5px; text-align: right; }
  .vazio { color: __TXT_VAZIO__; font-size: 12px; }
  .anuncio-linha { display: flex; align-items: center; gap: 10px; margin-bottom: 10px; font-size: 12.5px; }
  .anuncio-linha .nome { width: 150px; color: __TXT_NOME_LEAD__; flex-shrink: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .barra-fundo { flex: 1; height: 8px; background: __BARRA_FUNDO__; border-radius: 5px; overflow: hidden; }
  .barra-cheia { height: 100%; background: #D85A30; border-radius: 5px; }
  .anuncio-linha .qtd { width: 22px; text-align: right; color: __TXT_LBL__; }
  .rodape { text-align: center; color: __TXT_RODAPE__; font-size: 11px; padding: 18px 0 0; }
  @media (max-width: 760px) { .kpis { grid-template-columns: 1fr 1fr; } .linha { grid-template-columns: 1fr; } }
</style>
</head>
<body>
<div class="tela" id="tela-captura">
  <div class="nav">
    <div class="logo">__LOGO_SI__ Validador de Leads</div>
    <div class="dir">
      __BOTAO_EXCEL__
      <button class="baixar baixar-imagem" id="btnBaixarImagem" type="button">Baixar imagem</button>
    </div>
  </div>
  <div id="area-captura">
    <div class="hero">
      <div>
        <h1>__EMPRESA__</h1>
        <p>chave __CHAVE__ &nbsp;·&nbsp; __PERIODO__ &nbsp;·&nbsp; gerado em __GERADO__</p>
      </div>
      <div class="selo"><b>__TOTAL__</b> leads analisados</div>
    </div>

    <div class="corpo">
      <div class="kpis">
        <div class="kpi"><div class="lbl">Total de leads</div><div class="num">__TOTAL__</div><div class="sub">no período</div></div>
        <div class="kpi"><div class="lbl">Dentro do foco</div><div class="num v">__PCT_DENTRO__%</div><div class="sub">__N_DENTRO__ leads</div></div>
        <div class="kpi"><div class="lbl">Fora do foco</div><div class="num r">__PCT_FORA__%</div><div class="sub">__N_FORA__ leads</div></div>
        <div class="kpi"><div class="lbl">Aberto</div><div class="num a">__PCT_ABERTO__%</div><div class="sub">__N_ABERTO__ leads</div></div>
      </div>

      <div class="linha">
        <div class="painel">
          <h2>Distribuição por status</h2>
          <div class="grafico"><canvas id="rosca"></canvas></div>
        </div>
        <div class="painel">
          <h2>Anúncios que mais geraram leads fora do foco</h2>
          __LINHAS_ANUNCIOS__
        </div>
      </div>

      <div class="linha">
        <div class="painel">
          <h2>Melhores leads <span class="pin pin-verde">dentro do foco</span></h2>
          __LINHAS_MELHORES__
        </div>
        <div class="painel">
          <h2>Piores leads <span class="pin pin-verm">fora do foco</span></h2>
          __LINHAS_PIORES__
        </div>
      </div>
    </div>
  </div>
</div>
<p class="rodape">Validador de Leads · Soluções Industriais · uso interno</p>
<script>
new Chart(document.getElementById("rosca"), {
  type: "doughnut",
  data: {
    labels: ["Dentro do foco", "Fora do foco", "Aberto"],
    datasets: [{ data: [__N_DENTRO__, __N_FORA__, __N_ABERTO__], backgroundColor: ["#1D9E75", "#D85A30", "#BA7517"], borderWidth: 0 }]
  },
  options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: "bottom", labels: { color: "__COR_LEGENDA__", font: { size: 12 } } } } }
});

document.getElementById("btnBaixarImagem").addEventListener("click", function () {
  var btn = this;
  btn.textContent = "Gerando imagem...";
  html2canvas(document.getElementById("area-captura"), { backgroundColor: "__BG_TELA__", scale: 2, useCORS: true })
    .then(function (canvas) {
      var link = document.createElement("a");
      link.download = "__NOME_IMAGEM__.jpg";
      link.href = canvas.toDataURL("image/jpeg", 0.92);
      link.click();
      btn.textContent = "Baixar imagem";
    })
    .catch(function () {
      btn.textContent = "Erro ao gerar — tente de novo";
    });
});
</script>
</body>
</html>"""

# Cores do dashboard por tema — o dashboard baixado nasce fixo na cor do tema
# que estava ativo no app no momento da validação (não tem alternador nele).
TEMAS_DASH = {
    "escuro": {
        "BG_PAGINA": "#10161F", "BG_TELA": "#0C1D30",
        "BG_NAV": "rgba(255,255,255,0.05)", "BORDA_NAV": "rgba(255,255,255,0.08)",
        "TXT_LOGO": "#ffffff",
        "HERO_OVERLAY": "linear-gradient(160deg, rgba(6,20,36,0.62), rgba(12,68,124,0.5))",
        "TXT_TITULO": "#ffffff", "TXT_SUB": "#D9EAFB",
        "SELO_BG": "rgba(255,255,255,0.16)", "SELO_BORDA": "rgba(255,255,255,0.32)", "SELO_TXT": "#ffffff",
        "BG_KPI": "rgba(255,255,255,0.08)", "BORDA_KPI": "rgba(255,255,255,0.14)",
        "TXT_LBL": "#AFCBE8", "TXT_NUM": "#ffffff",
        "TXT_NOME_LEAD": "#EAF3FC", "TXT_EMAIL_LEAD": "#7C93AC", "BORDA_LISTA": "rgba(255,255,255,0.08)",
        "TXT_VAZIO": "rgba(234,243,252,0.55)", "BARRA_FUNDO": "rgba(255,255,255,0.12)",
        "BAIXAR_BG": "#ffffff", "BAIXAR_TXT": "#0C447C",
        "TXT_RODAPE": "rgba(234,243,252,0.5)", "COR_LEGENDA": "#EAF3FC",
    },
    "claro": {
        "BG_PAGINA": "#EEF3FA", "BG_TELA": "#EEF3FA",
        "BG_NAV": "rgba(255,255,255,0.75)", "BORDA_NAV": "rgba(24,95,165,0.10)",
        "TXT_LOGO": "#0C2036",
        "HERO_OVERLAY": "linear-gradient(160deg, rgba(230,241,251,0.88), rgba(181,212,244,0.68))",
        "TXT_TITULO": "#0C2036", "TXT_SUB": "#33475C",
        "SELO_BG": "rgba(255,255,255,0.7)", "SELO_BORDA": "rgba(24,95,165,0.2)", "SELO_TXT": "#0C447C",
        "BG_KPI": "rgba(255,255,255,0.72)", "BORDA_KPI": "rgba(255,255,255,0.9)",
        "TXT_LBL": "#5C7089", "TXT_NUM": "#0C2036",
        "TXT_NOME_LEAD": "#0C2036", "TXT_EMAIL_LEAD": "#5C7089", "BORDA_LISTA": "rgba(24,95,165,0.12)",
        "TXT_VAZIO": "#7C8CA1", "BARRA_FUNDO": "rgba(24,95,165,0.12)",
        "BAIXAR_BG": "#0C447C", "BAIXAR_TXT": "#ffffff",
        "TXT_RODAPE": "#7C8CA1", "COR_LEGENDA": "#33475C",
    },
}


def gerar_xlsx(cabecalho, registros, leads, classificacoes):
    """Gera um Excel com as colunas originais + STATUS + MOTIVO, pintando cada
    linha por status: verde (dentro), vermelho (fora), âmbar (aberto)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads validados"
    colunas = cabecalho + ["STATUS", "MOTIVO"]
    ws.append(colunas)
    for cel in ws[1]:
        cel.fill = FILL_CABECALHO
        cel.font = Font(color="FFFFFF", bold=True)
    for i, r in enumerate(registros):
        c = classificacoes.get(leads[i]["id"], {
            "status": "Aberto", "motivo": "Não classificado pela IA — revisar manualmente.",
        })
        linha = list(r) + [c["status"], c["motivo"]]
        ws.append(linha)
        fill = {"Dentro do foco": FILL_DENTRO, "Fora do foco": FILL_FORA}.get(c["status"], FILL_ABERTO)
        for cel in ws[ws.max_row]:
            cel.fill = fill
    # larguras aproximadas para leitura
    for col in ws.columns:
        larg = min(60, max(12, max((len(str(c.value)) if c.value else 0) for c in col) + 2))
        ws.column_dimensions[col[0].column_letter].width = larg
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_dashboard_html(empresa, chave, periodo, total, contagem,
                         melhores=None, piores=None, anuncios_ruins=None,
                         tema="escuro", xlsx_bytes=None, xlsx_nome="leads.xlsx"):
    def pct(n):
        return str(round(100 * n / total)) if total else "0"

    def linhas_leads(lista, cor):
        if not lista:
            return "<p class='vazio'>Nenhum lead nesta categoria.</p>"
        out = ""
        for ld in lista:
            out += (f"<div class='lista-lead'><div class='nome'>"
                    f"<span class='pin-dot pin-dot-{cor}'></span>#{ld['id']} · {ld['nome']}</div>"
                    f"<div class='email'>{ld['email']}</div></div>")
        return out

    def linhas_anuncios(lista):
        if not lista:
            return "<p class='vazio'>Sem dados de anúncio.</p>"
        maior = max(qtd for _, qtd in lista) or 1
        out = ""
        for nome, qtd in lista:
            largura = round(100 * qtd / maior)
            out += (f"<div class='anuncio-linha'><div class='nome'>{nome}</div>"
                    f"<div class='barra-fundo'><div class='barra-cheia' style='width:{largura}%'></div></div>"
                    f"<div class='qtd'>{qtd}</div></div>")
        return out

    def botao_excel():
        if not xlsx_bytes:
            return ""
        b64 = base64.b64encode(xlsx_bytes).decode()
        return (
            f'<a class="baixar" download="{xlsx_nome}" '
            f'href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}">'
            f'Baixar Excel</a>'
        )

    cores = TEMAS_DASH.get(tema, TEMAS_DASH["escuro"])

    html = MODELO_DASH
    trocas = {
        "__EMPRESA__": empresa,
        "__CHAVE__": chave,
        "__PERIODO__": periodo,
        "__GERADO__": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "__TOTAL__": str(total),
        "__PCT_DENTRO__": pct(contagem["Dentro do foco"]),
        "__PCT_FORA__": pct(contagem["Fora do foco"]),
        "__PCT_ABERTO__": pct(contagem["Aberto"]),
        "__N_DENTRO__": str(contagem["Dentro do foco"]),
        "__N_FORA__": str(contagem["Fora do foco"]),
        "__N_ABERTO__": str(contagem["Aberto"]),
        "__LINHAS_MELHORES__": linhas_leads(melhores, "verde"),
        "__LINHAS_PIORES__": linhas_leads(piores, "verm"),
        "__LINHAS_ANUNCIOS__": linhas_anuncios(anuncios_ruins),
        "__BOTAO_EXCEL__": botao_excel(),
        "__NOME_IMAGEM__": re.sub(r"[^\w\-]+", "-", f"dashboard-{empresa}").strip("-").lower() or "dashboard",
        "__LOGO_SI__": svg_logo_si(cores["TXT_LOGO"], cores["BG_TELA"]),
    }
    for chave_cor, valor_cor in cores.items():
        trocas[f"__{chave_cor}__"] = valor_cor
    for k, v in trocas.items():
        html = html.replace(k, v)
    return html


# ---------- IA (Cerebras ou Groq, compatíveis com OpenAI) ----------

def provedores_ativos():
    """Retorna a lista de provedores com chave configurada (Cerebras primeiro)."""
    return [n for n in ("cerebras", "groq") if secret(PROVEDORES[n]["chave"])]


def _extrair_lista(parsed):
    if isinstance(parsed, list):
        return parsed
    for v in parsed.values():
        if isinstance(v, list):
            return v
    raise ValueError("Resposta sem lista de resultados.")


def _tentar_modelo(url, api_key, modelo, conteudo_user):
    """Uma chamada a um modelo específico. Retorna (lista, None) em sucesso,
    ('404', None) se o modelo não existe, ('cota', horas) se a cota diária estourou,
    (None, excecao) em outros erros."""
    corpo = {
        "model": modelo,
        "temperature": 0.1,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": PROMPT_SISTEMA},
            {"role": "user", "content": conteudo_user},
        ],
    }
    ultima = None
    for tentativa in range(1, MAX_TENTATIVAS + 1):
        try:
            r = requests.post(url, json=corpo, headers={"Authorization": f"Bearer {api_key}"}, timeout=120)
            if r.status_code in (400, 404):
                return "404", None                      # modelo indisponível: tenta o próximo
            if r.status_code == 429 or r.status_code >= 500:
                try:
                    espera = float(r.headers.get("retry-after", 0))
                except (TypeError, ValueError):
                    espera = 0
                if espera > 300:
                    return "cota", espera
                time.sleep(min(60, espera + 1) if espera else min(30, 5 * tentativa))
                continue
            r.raise_for_status()
            texto = r.json()["choices"][0]["message"]["content"]
            return _extrair_lista(json.loads(texto)), None
        except Exception as e:
            ultima = e
            time.sleep(2)
    return None, ultima


def chamar_ia(perfil, lote, ordem, modelo_forcado=None):
    """Tenta os provedores em ordem; dentro de cada um, tenta a lista de modelos candidatos
    (começando pelo escolhido). Pula modelos indisponíveis (404) e troca de provedor na cota."""
    leads_texto = "\n\n".join(
        f"LEAD id={l['id']}\nMensagem: {l['mensagem']}\nContexto extra: {l['extra']}"
        for l in lote
    )
    conteudo_user = f"PERFIL DO CLIENTE:\n{perfil}\n\nLEADS A CLASSIFICAR:\n{leads_texto}"
    ultima = None
    esgotados = []
    for i, nome in enumerate(ordem):
        cfg = PROVEDORES[nome]
        api_key = secret(cfg["chave"])
        # ordem de modelos: o escolhido (se for deste provedor) primeiro, depois os candidatos
        modelos = list(cfg["modelos"])
        if modelo_forcado and i == 0 and modelo_forcado in modelos:
            modelos.remove(modelo_forcado)
            modelos.insert(0, modelo_forcado)
        elif modelo_forcado and i == 0:
            modelos.insert(0, modelo_forcado)

        cota_estourou = False
        for modelo in modelos:
            resultado, err = _tentar_modelo(cfg["url"], api_key, modelo, conteudo_user)
            if isinstance(resultado, list):
                return resultado
            if resultado == "404":
                continue                                # modelo indisponível: próximo candidato
            if resultado == "cota":
                esgotados.append(nome)
                cota_estourou = True
                break
            ultima = err                                # erro genérico: tenta próximo modelo
        if cota_estourou:
            continue                                    # próximo provedor
    if esgotados and len(esgotados) == len(ordem):
        raise RuntimeError(
            "Cota diária esgotada em todos os provedores de IA configurados "
            f"({', '.join(esgotados)}). Renova às 21h (horário de Brasília), ou adicione "
            "outra chave (CEREBRAS_API_KEY ou GROQ_API_KEY) nos Secrets para continuar agora."
        )
    if ultima:
        raise ultima
    raise RuntimeError(
        "Nenhum modelo respondeu nos provedores configurados. "
        "Confira se as chaves de IA nos Secrets estão corretas e ativas."
    )


# ---------- Interface ----------

st.set_page_config(page_title="Validador de Leads v3", page_icon="✅", layout="wide")

FOTO_INDUSTRIA = "https://images.unsplash.com/photo-1513828583688-c52646db42da?w=1900&q=70"

# Tema visual: claro ou escuro, trocável a qualquer momento pelo alternador no topo.
# Fica só na sessão (não tenta "lembrar" entre acessos diferentes — mantém simples).
TEMAS_APP = {
    "escuro": {
        "fundo": "#0B141C",
        "fundo_grad": "radial-gradient(at 0% 0%, rgba(46,123,255,0.16) 0px, transparent 55%), "
                      "radial-gradient(at 100% 0%, rgba(0,240,255,0.10) 0px, transparent 55%)",
        "hero_overlay": "linear-gradient(90deg, rgba(11,20,28,0.94), rgba(11,20,28,0.55))",
        "label": "#EAF3FC",
        "texto": "#9AA5B8",
        "painel_bg": "rgba(24,32,40,0.45)",
        "painel_borda": "rgba(255,255,255,0.08)",
        "input_bg": "#141C24",
        "input_borda": "rgba(255,255,255,0.10)",
        "input_txt": "#ffffff",
        "placeholder": "#55698A",
        "popover_bg": "#101820",
        "popover_txt": "#EAF3FC",
        "sec_btn_bg": "transparent",
        "sec_btn_borda": "transparent",
        "sec_btn_txt": "#9AA5B8",
        "hero_titulo": "#ffffff",
        "hero_sub": "#B8C6DA",
        "badge_bg": "rgba(255,255,255,0.06)",
        "badge_borda": "rgba(255,255,255,0.12)",
        "badge_txt": "#ffffff",
        "icone": "#7C8CA3",
        "icone_bg": "rgba(46,123,255,0.14)",
    },
    "claro": {
        "fundo": "#EEF3FA",
        "fundo_grad": "radial-gradient(at 0% 0%, rgba(46,123,255,0.08) 0px, transparent 55%), "
                      "radial-gradient(at 100% 0%, rgba(0,180,255,0.06) 0px, transparent 55%)",
        "hero_overlay": "linear-gradient(90deg, rgba(230,241,251,0.95), rgba(181,212,244,0.55))",
        "label": "#0C2036",
        "texto": "#5C7089",
        "painel_bg": "rgba(255,255,255,0.72)",
        "painel_borda": "rgba(255,255,255,0.9)",
        "input_bg": "#ffffff",
        "input_borda": "rgba(24,95,165,0.16)",
        "input_txt": "#0C2036",
        "placeholder": "#8CA0B8",
        "popover_bg": "#FFFFFF",
        "popover_txt": "#0C2036",
        "sec_btn_bg": "transparent",
        "sec_btn_borda": "transparent",
        "sec_btn_txt": "#0C447C",
        "hero_titulo": "#0C2036",
        "hero_sub": "#33475C",
        "badge_bg": "rgba(255,255,255,0.7)",
        "badge_borda": "rgba(24,95,165,0.18)",
        "badge_txt": "#0C447C",
        "icone": "#7C93AC",
        "icone_bg": "rgba(46,123,255,0.12)",
    },
}

# Cores do gradiente neon — iguais nos dois temas (é a assinatura visual nova).
NEON_1, NEON_2 = "#2E7BFF", "#00CFFF"

tema = st.session_state.get("tema", "escuro")
T = TEMAS_APP[tema]

# Marca "Soluções Industriais" na cor certa para o tema atual do app.
LOGO_SI = svg_logo_si(T['badge_txt'], T['fundo'])

st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Geist:wght@400;500&family=JetBrains+Mono:wght@500&display=swap');
  @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:wght,FILL@100..700,0..1');

  html, body, [data-testid="stAppViewContainer"] {{
    background-color: {T['fundo']};
    background-image: {T['fundo_grad']};
  }}
  [data-testid="stHeader"] {{ background: transparent; }}

  .block-container {{ max-width: 1180px; padding-top: 1.4rem; padding-bottom: 3rem; font-family: 'Geist', sans-serif; }}

  .block-container label, .block-container [data-testid="stWidgetLabel"] p {{
    color: {T['label']} !important; font-weight: 600 !important; font-size: 0.82rem !important;
  }}
  .block-container p, .block-container .stMarkdown {{ color: {T['texto']}; }}

  .topbar-vidro {{ display: flex; align-items: center; justify-content: space-between; padding: 2px 4px 18px; }}
  .topbar-vidro .marca {{
    display: flex; align-items: center; gap: 10px; color: {T['label']};
    font-family: 'Sora', sans-serif; font-weight: 700; font-size: 14px;
  }}
  .topbar-vidro .marca .logo-si {{ flex-shrink: 0; }}

  .hero {{
    background: {T['hero_overlay']}, url('{FOTO_INDUSTRIA}') center/cover no-repeat;
    border-radius: 20px; padding: 34px 40px; margin-bottom: 20px;
  }}
  .hero .badge {{
    display: inline-flex; align-items: center; gap: 7px; font-family: 'JetBrains Mono', monospace;
    font-size: 0.7rem; font-weight: 500; letter-spacing: 0.02em;
    color: {T['badge_txt']}; background: {T['badge_bg']}; border: 1px solid {T['badge_borda']};
    padding: 4px 13px; border-radius: 999px; margin-bottom: 12px;
  }}
  .hero .badge::before {{
    content: ""; width: 6px; height: 6px; border-radius: 50%; background: {NEON_1}; flex-shrink: 0;
  }}
  .hero h1 {{
    color: {T['hero_titulo']}; font-family: 'Sora', sans-serif;
    font-size: clamp(1.5rem, 3vw, 2rem); font-weight: 700; margin: 0 0 8px;
  }}
  .hero p {{ color: {T['hero_sub']}; font-size: 0.95rem; line-height: 1.5; max-width: 540px; margin: 0; }}

  .painel-titulo {{ display: flex; align-items: center; gap: 12px; margin-bottom: 22px; }}
  .painel-titulo .icone-titulo {{
    width: 36px; height: 36px; border-radius: 10px; background: {T['icone_bg']}; color: {NEON_1};
    display: flex; align-items: center; justify-content: center; font-size: 20px; flex-shrink: 0;
  }}
  .painel-titulo h2 {{ font-family: 'Sora', sans-serif; font-size: 1.05rem; font-weight: 700; }}

  /* Painéis com borda (formulário e histórico) ganham a estética de vidro.
     Zeramos margem/sombra extra do próprio Streamlit para não sobrar aquela
     "moldura dentro da moldura" — só uma borda visível, a do painel mesmo. */
  div[data-testid="stVerticalBlockBorderWrapper"] {{
    background: {T['painel_bg']} !important;
    border: 1px solid {T['painel_borda']} !important;
    border-radius: 18px !important;
    backdrop-filter: blur(16px);
    margin: 0 !important;
    box-shadow: none !important;
  }}
  div[data-testid="stVerticalBlockBorderWrapper"] > div {{
    margin: 0 !important;
  }}
  /* Reforço específico no painel do formulário (chave "painel_form"): remove
     qualquer respiro/linha que o próprio Streamlit injete por dentro, pra não
     sobrar aquela sensação de "moldura dentro da moldura". */
  .st-key-painel_form {{
    margin: 0 !important;
    padding: 0 !important;
  }}
  .st-key-painel_form > div,
  .st-key-painel_form [data-testid="stVerticalBlock"] {{
    margin: 0 !important;
    gap: 1rem !important;
  }}

  /* Campos em vidro — cobrimos tanto o div[data-baseweb=...] (usado por alguns
     campos) quanto o wrapper direto do Streamlit (usado por outros), senão parte
     dos campos fica branca. Pinta todo mundo por dentro e deixa só uma borda fora. */
  div[data-testid="stTextInput"], div[data-testid="stTextArea"],
  div[data-testid="stDateInput"] > div, div[data-testid="stSelectbox"] > div {{
    border: 1px solid {T['input_borda']} !important;
    border-radius: 12px !important;
    overflow: hidden;
  }}
  div[data-testid="stTextInput"] div, div[data-testid="stTextArea"] div,
  div[data-testid="stDateInput"] div, div[data-testid="stSelectbox"] div,
  div[data-baseweb="input"], div[data-baseweb="base-input"],
  div[data-baseweb="textarea"], div[data-baseweb="select"] > div {{
    background: {T['input_bg']} !important;
  }}
  div[data-testid="stTextInput"] input, div[data-testid="stTextArea"] textarea,
  div[data-testid="stDateInput"] input, div[data-testid="stSelectbox"] *,
  div[data-baseweb="input"] input, div[data-baseweb="textarea"] textarea {{
    background: transparent !important; color: {T['input_txt']} !important;
  }}
  div[data-testid="stTextInput"] input::placeholder,
  div[data-testid="stTextArea"] textarea::placeholder {{
    color: {T['placeholder']} !important;
  }}
  div[data-testid="stTextInput"]:focus-within, div[data-testid="stTextArea"]:focus-within,
  div[data-baseweb="input"]:focus-within, div[data-baseweb="textarea"]:focus-within {{
    border-color: {NEON_1} !important; box-shadow: 0 0 0 3px rgba(46,123,255,0.25) !important;
  }}
  /* Ícone de ajuda (?) — some se não forçarmos a cor */
  [data-testid="stTooltipIcon"], [data-testid="stTooltipIcon"] svg,
  [data-testid="stTooltipHoverTarget"] svg {{
    color: {T['texto']} !important; fill: {T['texto']} !important; opacity: 1 !important;
  }}
  div[data-testid="stTooltipContent"] {{
    background: {T['popover_bg']} !important; color: {T['popover_txt']} !important;
    border: 1px solid rgba(120,140,170,0.25) !important; border-radius: 10px !important;
  }}
  /* Menus flutuantes (opções do selectbox e o calendário do date_input) — no baseweb
     essas caixas têm fundo branco embutido em vários níveis, então zeramos tudo por
     dentro e pintamos só o nível de fora, na cor do tema ativo. */
  div[data-baseweb="popover"] {{
    background: {T['popover_bg']} !important;
    border: 1px solid rgba(120,140,170,0.25) !important;
    border-radius: 14px !important;
  }}
  div[data-baseweb="popover"] * {{ background: transparent !important; color: {T['popover_txt']} !important; }}
  div[data-baseweb="popover"] li:hover,
  div[data-baseweb="popover"] [aria-selected="true"]:not([role="gridcell"]) {{
    background: rgba(120,140,170,0.18) !important;
  }}
  div[data-baseweb="calendar"] {{ background: {T['popover_bg']} !important; }}
  div[data-baseweb="calendar"] [role="gridcell"][aria-selected="true"] div {{
    background: {NEON_1} !important; border-radius: 50% !important; color: #ffffff !important;
  }}
  div[data-baseweb="calendar"] [role="gridcell"]:hover div {{ background: rgba(120,140,170,0.18) !important; }}

  /* Botões em pílula, com feedback de pressão */
  .stButton > button, .stDownloadButton > button {{
    border-radius: 999px !important;
    transition: transform 140ms ease-out, box-shadow 160ms ease-out;
  }}
  .stButton > button:active, .stDownloadButton > button:active {{ transform: scale(0.97); }}
  .stButton > button[kind="primary"] {{
    background: linear-gradient(90deg, {NEON_1}, {NEON_2}) !important; border: none !important;
    font-weight: 700 !important; box-shadow: 0 0 18px rgba(46,123,255,0.40) !important;
  }}
  .stButton > button[kind="primary"]:hover {{ box-shadow: 0 0 26px rgba(46,123,255,0.60) !important; }}
  .stButton > button[kind="primary"] p {{ color: #ffffff !important; }}
  .stDownloadButton > button {{
    background: rgba(255,255,255,0.94) !important; border: none !important; font-weight: 700 !important;
  }}
  .stDownloadButton > button p {{ color: {AZUL_ESCURO} !important; }}
  .stButton > button[kind="secondary"] {{
    background: {T['sec_btn_bg']} !important; border: 1px solid {T['sec_btn_borda']} !important;
  }}
  .stButton > button[kind="secondary"]:hover {{ background: {T['painel_bg']} !important; }}
  .stButton > button[kind="secondary"] p {{ color: {T['sec_btn_txt']} !important; }}

  /* Ícones dentro dos campos (visual só — não interferem no clique/digitação) */
  .st-key-f_chave, .st-key-f_site, .st-key-f_obs,
  .st-key-f_data_inicio, .st-key-f_data_fim {{ position: relative; }}
  .st-key-f_chave::after, .st-key-f_site::after, .st-key-f_obs::after,
  .st-key-f_data_inicio::after, .st-key-f_data_fim::after {{
    font-family: 'Material Symbols Outlined'; font-size: 19px; color: {T['icone']};
    position: absolute; right: 14px; bottom: 11px; pointer-events: none;
  }}
  .st-key-f_chave::after {{ content: "fingerprint"; }}
  .st-key-f_site::after {{ content: "language"; }}
  .st-key-f_obs::after {{ content: "sticky_note_2"; bottom: auto; top: 34px; }}
  .st-key-f_data_inicio::after {{ content: "calendar_today"; font-size: 16px; }}
  .st-key-f_data_fim::after {{ content: "event_available"; font-size: 16px; }}

  /* Métricas como cards de vidro */
  div[data-testid="stMetric"] {{
    background: {T['painel_bg']} !important;
    border: 1px solid {T['painel_borda']} !important;
    border-radius: 16px; padding: 16px 18px; backdrop-filter: blur(14px);
  }}
  div[data-testid="stMetric"] label {{ color: {T['texto']} !important; font-weight: 600 !important; }}
  div[data-testid="stMetricValue"] {{ color: {T['label']} !important; font-weight: 700; }}
  div[data-testid="stMetricDelta"] {{ color: {T['texto']} !important; }}
</style>

<div class="topbar-vidro">
  <div class="marca">{LOGO_SI} Soluções Industriais · Validador de Leads</div>
</div>

<div class="hero">
  <span class="badge">v3 · vidro industrial</span>
  <h1>Nova validação</h1>
  <p>Informe a chave única do cliente e o período — a ferramenta busca sozinha o briefing, os orçamentos e os anúncios no Metabase.</p>
</div>
""", unsafe_allow_html=True)

col_topo_vazio, col_tema_esc, col_tema_cla = st.columns([5, 1, 1])
with col_tema_esc:
    if st.button("🌙 Escuro", use_container_width=True, key="btn_tema_escuro",
                 type=("primary" if tema == "escuro" else "secondary")):
        st.session_state["tema"] = "escuro"
        st.rerun()
with col_tema_cla:
    if st.button("☀️ Claro", use_container_width=True, key="btn_tema_claro",
                 type=("primary" if tema == "claro" else "secondary")):
        st.session_state["tema"] = "claro"
        st.rerun()

CAMPOS_FORM = ("f_chave", "f_site", "f_obs", "f_modelo")
if st.session_state.pop("limpar_form", False):
    for k in CAMPOS_FORM:
        st.session_state.pop(k, None)
    st.session_state.pop("resultado", None)

with st.container(border=True, key="painel_form"):
    st.markdown(
        f"<div class='painel-titulo'>"
        f"<span class='icone-titulo material-symbols-outlined'>fact_check</span>"
        f"<h2 style='margin:0; color:{T['label']};'>Configurações de validação</h2>"
        f"</div>",
        unsafe_allow_html=True,
    )
    grid_a, grid_b = st.columns(2)
    with grid_a:
        chave_unica = st.text_input("Chave única do cliente", placeholder="Ex.: 12-34567-1", key="f_chave")
    with grid_b:
        st.markdown(
            f"<div style='font-weight:600; font-size:0.82rem; color:{T['label']}; margin-bottom:0.4rem;'>Período</div>",
            unsafe_allow_html=True,
        )
        p_a, p_b = st.columns(2)
        with p_a:
            data_inicio = st.date_input(
                "Início", value=date.today() - timedelta(days=90), format="DD/MM/YYYY",
                label_visibility="collapsed", key="f_data_inicio",
            )
        with p_b:
            data_fim = st.date_input(
                "Fim", value=date.today(), format="DD/MM/YYYY",
                label_visibility="collapsed", key="f_data_fim",
            )

    grid_c, grid_d = st.columns(2)
    with grid_c:
        site = st.text_input(
            "Site do cliente (opcional, mas importante quando não há briefing cadastrado)",
            placeholder="https://www.sitedocliente.com.br", key="f_site",
        )
    with grid_d:
        obs = st.text_area(
            "Observações (opcional, mas importante quando não há briefing cadastrado)",
            placeholder="Ex.: cliente só vende máquinas (serviço, assistência, aluguel e peças = fora do foco).",
            height=90, key="f_obs",
        )

    linha_final_a, linha_final_b, linha_final_c = st.columns([2, 1, 1])
    with linha_final_a:
        modelo_escolha = st.selectbox(
            "Modelo de IA",
            list(MODELOS_ESCOLHA.keys()),
            key="f_modelo",
            help="Rápido = resposta mais veloz, boa para o dia a dia. Preciso = mais lento, "
                 "melhor em casos sutis. Automático usa o provedor com maior cota disponível.",
        )
    with linha_final_b:
        st.markdown("<div style='height:1.85rem;'></div>", unsafe_allow_html=True)
        confirmando_limpeza = st.session_state.get("confirmar_limpar", False)
        rotulo_limpar = "Confirmar limpeza?" if confirmando_limpeza else "Limpar campos"
        if st.button(rotulo_limpar, use_container_width=True, key="btn_limpar",
                      type=("primary" if confirmando_limpeza else "secondary")):
            if confirmando_limpeza:
                st.session_state["confirmar_limpar"] = False
                st.session_state["limpar_form"] = True
                st.rerun()
            else:
                st.session_state["confirmar_limpar"] = True
                st.rerun()
    with linha_final_c:
        st.markdown("<div style='height:1.85rem;'></div>", unsafe_allow_html=True)
        botao_validar_ph = st.empty()
        with botao_validar_ph.container():
            validar = st.button("⚡ Validar leads", type="primary", use_container_width=True, key="btn_validar")

TOTAL_ETAPAS = 5


def marcar_etapa(n, texto):
    """Substitui a área do botão por um indicador de progresso ao vivo — o
    analista sempre sabe em que etapa está e quanto falta, e não dá pra
    clicar duas vezes por engano (o botão literalmente não existe mais ali)."""
    with botao_validar_ph.container():
        st.markdown(
            f"<div style='background: rgba(46,123,255,0.16); border: 1px solid rgba(46,123,255,0.45); "
            f"color: {T['label']}; text-align: center; padding: 11px 10px; border-radius: 999px; "
            f"font-size: 0.82rem; font-weight: 700;'>⏳ Etapa {n} de {TOTAL_ETAPAS} · {texto}</div>",
            unsafe_allow_html=True,
        )


def marcar_etapa_erro(n, texto):
    """Igual ao marcar_etapa, mas em vermelho — deixa claro que parou por erro
    na etapa N, e não que está apenas travado/lento."""
    with botao_validar_ph.container():
        st.markdown(
            f"<div style='background: rgba(216,90,48,0.16); border: 1px solid rgba(216,90,48,0.55); "
            f"color: #D85A30; text-align: center; padding: 11px 10px; border-radius: 999px; "
            f"font-size: 0.82rem; font-weight: 700;'>⚠️ Erro na Etapa {n} · {texto}</div>",
            unsafe_allow_html=True,
        )


def montar_regras():
    if obs.strip():
        return f"- Observações do projeto (prioridade máxima): {obs.strip()}"
    return ""

if validar:
    st.session_state.pop("resultado", None)  # some o resultado da consulta anterior — evita baixar o arquivo errado
    st.session_state["confirmar_limpar"] = False  # desarma um "Confirmar limpeza?" pendente
    ordem_ia = provedores_ativos()
    if not ordem_ia:
        st.error("Nenhuma chave de IA configurada. Adicione CEREBRAS_API_KEY ou GROQ_API_KEY nos Secrets.")
        st.stop()
    if not secret("METABASE_URL"):
        st.error("Segredo METABASE_URL não configurado (ex.: https://metabase.ferramentademarketing.com.br).")
        st.stop()
    if not chave_unica.strip():
        st.error("Preencha a chave única do cliente.")
        st.stop()
    if data_inicio > data_fim:
        st.error("A data de início não pode ser depois da data de fim.")
        st.stop()

    # Traduz a escolha do seletor em ordem de provedores + modelo fixo (se houver)
    modelo_forcado = None
    escolha = MODELOS_ESCOLHA.get(modelo_escolha)
    if escolha:
        prov, modelo_forcado = escolha
        if prov not in ordem_ia:
            st.error(f"O modelo escolhido usa {prov.title()}, mas não há chave desse provedor nos Secrets. "
                     "Escolha outro modelo ou adicione a chave.")
            st.stop()
        ordem_ia = [prov] + [p for p in ordem_ia if p != prov]   # escolhido primeiro, resto como reserva

    nomes_ia = {"cerebras": "Cerebras", "groq": "Groq"}
    if modelo_forcado:
        st.caption(f"IA: {modelo_escolha}"
                   + (f" · reserva: {nomes_ia[ordem_ia[1]]}" if len(ordem_ia) > 1 else ""))
    else:
        st.caption("IA: " + " → ".join(nomes_ia[n] for n in ordem_ia)
                   + (" (reserva automática)" if len(ordem_ia) > 1 else ""))

    # 1. Briefing (question 286 — por chave única). Nem todo cliente tem briefing cadastrado.
    marcar_etapa(1, "Buscando briefing no Metabase...")
    try:
        csv_briefing = consultar_question(CARD_BRIEFING, [
            ("chave_unica", chave_unica.strip(), "category"),
        ])
    except Exception as e:
        marcar_etapa_erro(1, "falha ao buscar briefing")
        st.error(f"Erro ao buscar o briefing (question {CARD_BRIEFING}): {e}")
        st.stop()
    texto_briefing = csv_briefing_para_texto(csv_briefing)
    briefing_ausente = not texto_briefing
    if briefing_ausente:
        st.warning(
            f'Nenhum briefing cadastrado para a chave "{chave_unica}". '
            "A IA vai se basear no site informado e nas observações do projeto — "
            "preencha ao menos um desses dois campos para esse cliente."
        )

    nome_empresa = extrair_nome_empresa(csv_briefing) or chave_unica.strip()

    # 2. Orçamentos (question 47)
    marcar_etapa(2, "Buscando orçamentos no Metabase...")
    try:
        csv_orcamentos = consultar_question(CARD_ORCAMENTOS, [
            ("chave_unica", chave_unica.strip(), "category"),
            ("data_inicio", data_inicio.isoformat(), "date/single"),
            ("data_fim", data_fim.isoformat(), "date/single"),
        ])
    except Exception as e:
        marcar_etapa_erro(2, "falha ao buscar orçamentos")
        st.error(f"Erro ao buscar os orçamentos (question {CARD_ORCAMENTOS}): {e}")
        st.stop()
    linhas = list(csv.reader(io.StringIO(csv_orcamentos)))
    if len(linhas) < 2:
        marcar_etapa_erro(2, "nenhum orçamento encontrado")
        st.error("Nenhum orçamento encontrado para essa chave única nesse período.")
        st.stop()

    # Se o briefing não trouxe o nome, tenta pelos orçamentos (coluna "Nome Fantasia")
    if nome_empresa == chave_unica.strip():
        nome_empresa = extrair_nome_empresa(csv_orcamentos) or nome_empresa
    st.caption(f"Cliente identificado: {nome_empresa}")

    # 3. Anúncios ativos (question 185 — opcional, não bloqueia se falhar)
    texto_anuncios = ""
    marcar_etapa(3, "Buscando anúncios do cliente...")
    try:
        csv_anuncios = consultar_question(CARD_ANUNCIOS, [
            ("chave_unica", chave_unica.strip(), "category"),
        ])
        texto_anuncios = csv_anuncios_para_texto(csv_anuncios)
    except Exception:
        st.warning("Não consegui buscar os anúncios (question 185) — prosseguindo sem eles.")

    # 4. Site do cliente (opcional — essencial quando não há briefing)
    texto_site = ""
    marcar_etapa(4, "Preparando perfil do cliente...")
    if site.strip():
        texto_site = buscar_site(site.strip())
        if not texto_site:
            st.warning("Não consegui ler o site informado — prosseguindo sem ele.")

    # 5. Perfil do cliente — monta com o que houver disponível (briefing e/ou site e/ou observações)
    partes_perfil = []
    regras_projeto = montar_regras()
    if regras_projeto:
        partes_perfil.append(f"===== OBSERVAÇÕES DO PROJETO (prioridade máxima) =====\n{regras_projeto}")
    if texto_briefing:
        partes_perfil.append(f"===== BRIEFING DO CLIENTE (Metabase) =====\n{texto_briefing}")
    if texto_site:
        partes_perfil.append(f"===== SITE DO CLIENTE ({site.strip()}) =====\n{texto_site}")
    if texto_anuncios:
        partes_perfil.append(f"===== ANÚNCIOS ATIVOS DO CLIENTE (termos anunciados) =====\n{texto_anuncios}")
    perfil = "\n\n".join(partes_perfil)

    if not perfil.strip():
        marcar_etapa_erro(4, "perfil do cliente insuficiente")
        st.error(
            "Não há briefing, site nem observações suficientes para avaliar esse cliente. "
            "Preencha o site ou as observações do projeto e envie de novo."
        )
        st.stop()
    if len(perfil) > LIMITE_PERFIL:
        perfil = perfil[:LIMITE_PERFIL] + "\n[... perfil truncado para caber no limite da IA gratuita ...]"

    # 5. Montagem dos leads
    cabecalho = [c.strip() for c in linhas[0]]
    col_msg = "Mensagem do Cliente"
    if col_msg not in cabecalho:
        marcar_etapa_erro(4, "coluna de mensagem não encontrada")
        st.error(f'Coluna "{col_msg}" não encontrada no retorno do Metabase. Colunas: {", ".join(cabecalho)}')
        st.stop()
    idx_msg = cabecalho.index(col_msg)

    def idx_de(*nomes):
        for n in nomes:
            if n in cabecalho:
                return cabecalho.index(n)
        return -1

    idx_id = idx_de("ID do Orçamento", "ID do Orcamento")
    idx_nome = idx_de("Nome do Comprador", "Nome do comprador")
    idx_email = idx_de("E-mail do Comprador", "Email do Comprador", "E-mail do comprador")
    idx_anuncio = idx_de("anúncio de origem do Orçamento", "Anúncio do cliente", "anuncio de origem do Orçamento")

    registros = [r for r in linhas[1:] if any(c.strip() for c in r)]
    st.info(f"{len(registros)} leads encontrados de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}.")

    def celula(r, idx):
        return r[idx].strip() if 0 <= idx < len(r) else ""

    leads = []
    for i, r in enumerate(registros):
        extra = "; ".join(
            f"{cabecalho[j]}: {r[j]}" for j in range(len(cabecalho))
            if j != idx_msg and j < len(r) and r[j].strip()
        )[:500]
        leads.append({
            "id": f"L{i+2}",
            "mensagem": r[idx_msg] if idx_msg < len(r) else "",
            "extra": extra,
            "id_orc": celula(r, idx_id) or f"linha {i+2}",
            "nome": celula(r, idx_nome) or "(sem nome)",
            "email": celula(r, idx_email) or "(sem e-mail)",
            "anuncio": celula(r, idx_anuncio) or "(sem anúncio)",
            "tam_msg": len(celula(r, idx_msg)),
        })

    # 6. Classificação com re-tentativas
    classificacoes = {}
    erros_ia = []

    def processar(lista, tamanho_lote, rotulo):
        total_lotes = (len(lista) + tamanho_lote - 1) // tamanho_lote
        progresso = st.progress(0, text=f"{rotulo}: {len(lista)} leads...")
        for n in range(total_lotes):
            lote = lista[n * tamanho_lote:(n + 1) * tamanho_lote]
            try:
                resultado = chamar_ia(perfil, lote, ordem_ia, modelo_forcado)
            except RuntimeError as e:
                progresso.empty()
                marcar_etapa_erro(5, "cota de IA esgotada")
                st.error(str(e))       # cota diária esgotada: para tudo na hora
                st.stop()
            except Exception as e:
                erros_ia.append(f"{type(e).__name__}: {e}")
                resultado = []
            for item in resultado:
                status = str(item.get("status", "")).strip()
                if status not in STATUS_VALIDOS:
                    status = "Aberto"
                classificacoes[str(item.get("id", "")).strip()] = {
                    "status": status, "motivo": str(item.get("motivo", "")).strip(),
                }
            progresso.progress((n + 1) / total_lotes, text=f"{rotulo}: lote {n+1} de {total_lotes}")
            if n + 1 < total_lotes:
                time.sleep(1)
        progresso.empty()

    marcar_etapa(5, "Classificando leads com IA...")
    processar(leads, TAMANHO_LOTE, "Classificando")
    pendentes = [l for l in leads if l["id"] not in classificacoes]
    if pendentes:
        st.info(f"{len(pendentes)} lead(s) sem resposta — reprocessando em lotes menores...")
        time.sleep(5)
        processar(pendentes, 5, "Reprocessando")
    pendentes = [l for l in leads if l["id"] not in classificacoes]
    if pendentes:
        time.sleep(5)
        processar(pendentes, 1, "Última passada")
    falhas = sum(1 for l in leads if l["id"] not in classificacoes)

    # 7. Contagem + rankings para o dashboard
    contagem = {"Dentro do foco": 0, "Fora do foco": 0, "Aberto": 0}
    for i in range(len(registros)):
        c = classificacoes.get(leads[i]["id"])
        status = c["status"] if c else "Aberto"
        contagem[status] += 1
        leads[i]["status"] = status

    dentro = [l for l in leads if l.get("status") == "Dentro do foco"]
    fora = [l for l in leads if l.get("status") == "Fora do foco"]
    # melhores = dentro do foco com mensagem mais específica (mais longa) primeiro
    melhores = sorted(dentro, key=lambda l: l["tam_msg"], reverse=True)[:10]
    melhores = [{"id": l["id_orc"], "nome": l["nome"], "email": l["email"]} for l in melhores]
    piores = fora[:10]
    piores = [{"id": l["id_orc"], "nome": l["nome"], "email": l["email"]} for l in piores]
    # anúncios que mais geraram leads fora do foco
    cont_anuncios = {}
    for l in fora:
        cont_anuncios[l["anuncio"]] = cont_anuncios.get(l["anuncio"], 0) + 1
    anuncios_ruins = sorted(cont_anuncios.items(), key=lambda x: x[1], reverse=True)[:8]

    # 8. Excel com destaque + dashboard
    xlsx_bytes = gerar_xlsx(cabecalho, registros, leads, classificacoes)

    total = len(registros)
    periodo_txt = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    base_nome = f"{nome_empresa} - {data_inicio.isoformat()} a {data_fim.isoformat()}"
    dash_html = gerar_dashboard_html(nome_empresa, chave_unica.strip(), periodo_txt, total, contagem,
                                     melhores=melhores, piores=piores, anuncios_ruins=anuncios_ruins,
                                     tema=tema, xlsx_bytes=xlsx_bytes,
                                     xlsx_nome=f"{base_nome} - Validado.xlsx")

    # Resultado fica guardado na sessão: os downloads não somem ao clicar
    st.session_state["resultado"] = {
        "empresa": nome_empresa,
        "total": total,
        "contagem": contagem,
        "falhas": falhas,
        "erro_ia": erros_ia[-1] if erros_ia else "",
        "xlsx_bytes": xlsx_bytes,
        "xlsx_nome": f"{base_nome} - Validado.xlsx",
        "dash_bytes": dash_html.encode("utf-8"),
        "dash_nome": f"{base_nome} - Dashboard.html",
    }

    salvar_no_historico({
        "id": datetime.now().strftime("%Y%m%d%H%M%S"),
        "Data da solicitação": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Empresa": nome_empresa,
        "Chave única": chave_unica.strip(),
        "Período": periodo_txt,
        "Leads": total,
        "Dentro do foco": contagem["Dentro do foco"],
        "Fora do foco": contagem["Fora do foco"],
        "Aberto": contagem["Aberto"],
        "xlsx_nome": f"{base_nome} - Validado.xlsx",
        "dash_nome": f"{base_nome} - Dashboard.html",
    }, xlsx_bytes=xlsx_bytes,
       dash_bytes=st.session_state["resultado"]["dash_bytes"])

    botao_validar_ph.empty()

res = st.session_state.get("resultado")
if res:
    total, contagem = res["total"], res["contagem"]
    st.success(f"Validação concluída — {total} leads processados! ({res['empresa']})")
    c1, c2, c3 = st.columns(3)
    c1.metric("Dentro do foco", f"{contagem['Dentro do foco']/total:.0%}", f"{contagem['Dentro do foco']} leads", delta_color="off")
    c2.metric("Fora do foco", f"{contagem['Fora do foco']/total:.0%}", f"{contagem['Fora do foco']} leads", delta_color="off")
    c3.metric("Aberto", f"{contagem['Aberto']/total:.0%}", f"{contagem['Aberto']} leads", delta_color="off")
    if res["falhas"]:
        st.warning(f"{res['falhas']} lead(s) sem resposta da IA (marcados como Aberto) — rode de novo para reprocessar.")
        if res["erro_ia"]:
            st.error(f"Motivo técnico da falha na IA: {res['erro_ia'][:400]}")

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button("Baixar Excel validado", data=res["xlsx_bytes"],
                           file_name=res["xlsx_nome"],
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="dl_xlsx")
    with col_dl2:
        st.download_button("Baixar dashboard (HTML)", data=res["dash_bytes"],
                           file_name=res["dash_nome"], mime="text/html",
                           use_container_width=True, key="dl_dash")

@st.dialog("Excluir pesquisa")
def dialogo_excluir(rid, rotulo):
    st.write(f"Tem certeza que deseja excluir a pesquisa **{rotulo}**?")
    st.caption("O registro e os arquivos (CSV e dashboard) dela serão apagados. Essa ação não tem volta.")
    cd1, cd2 = st.columns(2)
    if cd1.button("Sim, excluir", type="primary", use_container_width=True, key=f"conf_{rid}"):
        excluir_do_historico(rid)
        st.rerun()
    if cd2.button("Cancelar", use_container_width=True, key=f"canc_{rid}"):
        st.rerun()


st.markdown("<p style='font-weight:600; margin-top:32px;'>Histórico de validações</p>", unsafe_allow_html=True)
historico = carregar_historico()
if historico:
    filtro_chave = st.text_input(
        "Buscar por chave única",
        placeholder="Digite a chave para filtrar o histórico (ex.: 12-34567-1)",
        key="f_filtro_hist",
    )
    if filtro_chave.strip():
        alvo = filtro_chave.strip().lower()
        historico = [h for h in historico if alvo in str(h.get("Chave única", "")).lower()]
        if not historico:
            st.caption("Nenhuma validação encontrada para essa chave.")
    colunas_hist = [1.3, 1.4, 1.1, 1.6, 1.0, 0.4, 0.4, 0.4]
    cab = st.columns(colunas_hist)
    for col, titulo in zip(cab, ("Data", "Empresa", "Chave", "Período", "Leads (D/F/A)", "", "", "")):
        col.markdown(f"<span style='font-size:0.72rem; color:{T['texto']}; font-weight:600;'>{titulo}</span>", unsafe_allow_html=True)

    for h in historico:
        rid = h.get("id", "")
        with st.container(border=True):
            c = st.columns(colunas_hist)
            c[0].markdown(f"<span style='font-size:0.78rem;'>{h.get('Data da solicitação', '')}</span>", unsafe_allow_html=True)
            c[1].markdown(f"<span style='font-size:0.78rem;'>{h.get('Empresa', '')}</span>", unsafe_allow_html=True)
            c[2].markdown(f"<span style='font-size:0.78rem;'>{h.get('Chave única', '')}</span>", unsafe_allow_html=True)
            c[3].markdown(f"<span style='font-size:0.78rem;'>{h.get('Período', '')}</span>", unsafe_allow_html=True)
            c[4].markdown(
                f"<span style='font-size:0.78rem;'>{h.get('Leads', '')} "
                f"(<span style='color:#22883A;'>{h.get('Dentro do foco', '')}</span>/"
                f"<span style='color:#D6433F;'>{h.get('Fora do foco', '')}</span>/"
                f"<span style='color:#B4830A;'>{h.get('Aberto', '')}</span>)</span>",
                unsafe_allow_html=True,
            )
            # download direto na própria linha — sem precisar abrir um seletor à parte
            xlsx_salvo = ler_resultado_salvo(rid, ".xlsx") if rid else None
            dash_salvo = ler_resultado_salvo(rid, ".html") if rid else None
            if xlsx_salvo:
                c[5].download_button("⬇︎", data=xlsx_salvo, file_name=h.get("xlsx_nome", f"{rid}.xlsx"),
                                      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                      use_container_width=True, key=f"hxlsx_{rid}", help="Baixar Excel")
            else:
                c[5].caption("—")
            if dash_salvo:
                c[6].download_button("🖥", data=dash_salvo, file_name=h.get("dash_nome", f"{rid}.html"),
                                      mime="text/html", use_container_width=True, key=f"hdash_{rid}",
                                      help="Baixar dashboard")
            else:
                c[6].caption("—")
            if rid and c[7].button("✕", key=f"x_{rid}", help="Excluir esta pesquisa"):
                dialogo_excluir(rid, f"{h.get('Empresa', '')} · {h.get('Data da solicitação', '')}")
else:
    st.caption("Nenhuma validação registrada ainda. As próximas aparecerão aqui com data, empresa e resultado.")

st.markdown(
    f"<p style='text-align:center; color:{T['texto']}; font-size:0.75rem; margin-top:32px;'>"
    "Validador de Leads v3 · Soluções Industriais</p>",
    unsafe_allow_html=True,
)
